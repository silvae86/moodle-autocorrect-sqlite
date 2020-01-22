# coding=utf8
#
# Moodle SQL Batch corrector by João Rocha da Silva (silvae86 on GitHub)
# Copyright 2020 João Rocha da Silva
# ==============================================================================


# Usage Example: "python batch_sql_corrector.py correction/cscript-students.sql correction/script-correction.sql correction/student_answers.xlsx "Resposta 16.sql"

from openpyxl.reader.excel import load_workbook
import os
import sys
import tempfile
import subprocess
import tableformatter as tf
import time
import codecs
import unicodedata

def find_column_with_first_row_equal_to(worksheet,value):
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == value:
                return cell.column
    return 0

def remove_control_characters(s):
    return "".join(ch for ch in s if unicodedata.category(ch)[0]!="C" or ch == '\n').replace(u"\u00A0", u" ")

def runcommand(cmd):
    ps = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    output = ps.communicate()[0]
    return str(output, encoding="utf8")

# indexes of the columns that have the corresponding information. to adjust, have into account that the first column is index 1 (one)
firstname_column = 2
surname_column = 1
email_column = 4

# row in the excel file where the answers start
start_row=2

# the path of the sql script given to the students (sample)
database_script_path_sample = str(sys.argv[1])

# the path of the full sql script (correction)
database_script_path = str(sys.argv[2])

# the path to the worksheet with the student answers from Moodle
all_student_answers = load_workbook(str(sys.argv[3]))

# get active worksheet
ws = all_student_answers.active

# the name of the column that contains the answers for the question that you want to evaluate    
question_name = str(sys.argv[4])
question_column = find_column_with_first_row_equal_to(ws, question_name)

if question_column <= 0 or question_column is None:
    print("Invalid question name: " + question_name + ". Is there a column named " + question_name + " in the Excel file with the students' answers?")
    quit()

# the path to the file containing the text of the proposed answer
proposed_answer_location = str(sys.argv[5])

summary_rows = [
    ("Date", time.strftime("%a, %d %b %Y %H:%M:%S")),
    ("SQLite Version", runcommand("sqlite3 --version")),
    ("Sample database script", database_script_path_sample),
    ("Complete database script", database_script_path)
]
print(tf.generate_table(summary_rows, grid_style=tf.FancyGrid()))

StatesPriority = {}
currentRow = 2

with open(proposed_answer_location, 'r') as f:
    proposed_answer = f.read()

for row in ws.iter_rows(min_row=start_row):
    firstname = ws.cell(row=currentRow, column=firstname_column).value
    surname = ws.cell(row=currentRow, column=surname_column).value
    email = ws.cell(row=currentRow, column=email_column).value

    # remove non breaking spaces
    student_answer = ws.cell(row=currentRow, column=question_column).value
    if student_answer is not None:
        student_answer_filtered = remove_control_characters(student_answer)
    else:
        student_answer_filtered = ""

    # Run student solution against the databases
    with tempfile.NamedTemporaryFile(mode='w+', encoding='utf-8') as f:
        if student_answer_filtered is not None:
            f.write(student_answer_filtered)

        f.flush()
        student_result_sample = runcommand("cat \"" + database_script_path_sample + "\" \"" + f.name + "\" | sqlite3")
        student_result = runcommand("cat \"" + database_script_path + "\" \"" + f.name + "\" | sqlite3")

    # Run proposed solutions against the databases
    proposed_result_sample = runcommand("cat \"" + database_script_path_sample + "\" \"" + proposed_answer_location + "\" | sqlite3")
    proposed_result = runcommand("cat \"" + database_script_path + "\" \"" + proposed_answer_location + "\" | sqlite3")

    rows = [
        ("SQL", student_answer, proposed_answer, ""),
        ("Example DB", student_result_sample, proposed_result_sample, student_result_sample == proposed_result_sample),
        ("Complete DB", student_result, proposed_result, student_result == proposed_result)
    ]

    columns = (
                tf.Column(''),
                tf.Column("Student: " + firstname + " " + surname + " " + email, width=80,  wrap_mode=tf.WrapMode.WRAP),
                tf.Column("Professor", width=80,  wrap_mode=tf.WrapMode.WRAP),
                tf.Column("Match")
               )

    print(tf.generate_table(rows, columns, grid_style=tf.FancyGrid()))


    currentRow = currentRow + 1

print("Results for " + question_name + " done.")
